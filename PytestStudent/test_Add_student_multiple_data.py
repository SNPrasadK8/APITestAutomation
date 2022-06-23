import requests
import json
import jsonpath
import openpyxl


def test_add_multiple():

    App_Url ="http://localhost:3000/studentDetails"
    f = open('C:/SNP/PyTestUdemy/PytestStudent/RequestJason.json', 'r')
    json_request = json.loads(f.read())

    vk = openpyxl.load_workbook('C:\SNP\PyTestUdemy\PytestStudent\Book1.xlsx')
    sh = vk['Sheet1']
    rows = sh.max_row
    
    for i in range(2, rows+1):
        cell_first_name  = sh.cell(row = i, column =1)
        cell_mid_name  = sh.cell(row = i, column =2)
        cell_last_name  = sh.cell(row = i, column =3)
        cell_dob  = sh.cell(row = i, column =4)
        
        json_request['first_name']=cell_first_name.value
        json_request['middle_name']=cell_mid_name.value
        json_request['last_name']=cell_last_name.value
        json_request['date_of_birth']=cell_dob.value

        response = requests.post(App_Url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201